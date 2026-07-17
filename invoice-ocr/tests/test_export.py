import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

SCRIPTS = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(SCRIPTS))

from export import PRICE_ALERT_FILL, export_xlsx_append


def make_table(path):
    workbook = Workbook()
    workbook.active.append([
        "供货单位",
        "货品编号",
        "结算价",
        "到货量",
        "单据号",
    ])
    workbook.save(path)


def batch(note="DN-1", price=10.0):
    return {
        "results": [{
            "filename": f"{note}.jpg",
            "status": "success",
            "review_status": "confirmed",
            "data": {
                "delivery_note": {
                    "supplier_name": "旺泰",
                    "note_number": note,
                    "date": "2026-06-25",
                },
                "items": [{
                    "fabric_code": "A100",
                    "unit_price": price,
                    "quantity": 20,
                    "unit": "米",
                    "total_amount": price * 20,
                }],
                "needs_review": [],
            },
        }],
    }


MAPPINGS = {
    1: {"header": "供货单位", "target_field": "supplier_name"},
    2: {"header": "货品编号", "target_field": "fabric_code"},
    3: {"header": "结算价", "target_field": "unit_price"},
    4: {"header": "到货量", "target_field": "quantity"},
    5: {"header": "单据号", "target_field": "note_number"},
}


class ExportTests(unittest.TestCase):
    def test_writes_to_custom_headers(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "table.xlsx"
            make_table(path)

            stats = export_xlsx_append(batch(), str(path), MAPPINGS)

            sheet = load_workbook(path).active
            self.assertEqual(sheet.cell(2, 1).value, "旺泰")
            self.assertEqual(sheet.cell(2, 3).value, 10.0)
            self.assertEqual(stats.written_rows, 1)

    def test_duplicate_note_is_skipped(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "table.xlsx"
            make_table(path)
            export_xlsx_append(
                batch("DN-1", 10.0),
                str(path),
                MAPPINGS,
            )

            stats = export_xlsx_append(
                batch("DN-1", 99.0),
                str(path),
                MAPPINGS,
            )

            sheet = load_workbook(path).active
            self.assertEqual(sheet.max_row, 2)
            self.assertEqual(sheet.cell(2, 3).value, 10.0)
            self.assertEqual(stats.skipped_duplicates, 1)

    def test_price_change_remains_red(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "table.xlsx"
            make_table(path)
            export_xlsx_append(
                batch("DN-1", 10.0),
                str(path),
                MAPPINGS,
            )

            stats = export_xlsx_append(
                batch("DN-2", 11.0),
                str(path),
                MAPPINGS,
            )

            sheet = load_workbook(path).active
            self.assertEqual(stats.price_alerts, 1)
            self.assertEqual(stats.price_alert_details[0]["supplier_name"], "旺泰")
            self.assertEqual(stats.price_alert_details[0]["fabric_code"], "A100")
            self.assertEqual(stats.price_alert_details[0]["previous_price"], 10.0)
            self.assertEqual(stats.price_alert_details[0]["current_price"], 11.0)
            self.assertEqual(
                sheet.cell(3, 3).fill.start_color.rgb,
                PRICE_ALERT_FILL.start_color.rgb,
            )

    def test_formula_like_text_is_written_as_text(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "table.xlsx"
            make_table(path)
            payload = batch()
            payload["results"][0]["data"]["delivery_note"][
                "supplier_name"
            ] = '=HYPERLINK("https://example.invalid")'

            export_xlsx_append(payload, str(path), MAPPINGS)

            value = load_workbook(path).active.cell(2, 1).value
            self.assertTrue(value.startswith("'="))


if __name__ == "__main__":
    unittest.main()
