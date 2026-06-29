import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from openpyxl import load_workbook

SCRIPTS = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(SCRIPTS))

import process


def workbook(path, headers):
    wb = Workbook()
    wb.active.append(headers)
    wb.save(path)


def valid_invoice_result(filename):
    return {
        "filename": filename,
        "status": "success",
        "review_status": "pending",
        "data": {
            "document_type": "delivery",
            "document_title": "销售码单",
            "delivery_note": {
                "supplier_name": "旺泰纺织",
                "note_number": "WT-001",
                "date": "2026-06-25",
            },
            "items": [{
                "material_type": "面料",
                "fabric_code": "A100",
                "unit_price": 10.0,
                "quantity": 20.0,
                "unit": "米",
                "total_amount": 200.0,
            }],
            "total_amount": 200.0,
            "needs_review": [],
        },
    }


class TargetTableTests(unittest.TestCase):
    def test_missing_table_waits_without_calling_llm(self):
        with patch.object(process, "infer_headers") as infer:
            result = process.prepare_target_table(None, None, {})

        self.assertEqual(result["status"], "waiting_for_table")
        infer.assert_not_called()

    def test_exact_headers_are_ready_without_llm(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "table.xlsx"
            workbook(path, ["单号", "单价", "数量"])
            with patch.object(process, "infer_headers") as infer:
                result = process.prepare_target_table(
                    str(path),
                    None,
                    {"base_url": "", "token": "", "model": ""},
                )

        self.assertEqual(result["status"], "ready")
        infer.assert_not_called()

    def test_same_signature_reuses_learned_mapping_without_llm(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "second.xlsx"
            workbook(path, ["供货单位", "结算价"])
            learned = {
                "mappings": {
                    "1": {
                        "header": "供货单位",
                        "target_field": "supplier_name",
                        "source": "user",
                        "confidence": 1.0,
                    },
                    "2": {
                        "header": "结算价",
                        "target_field": "unit_price",
                        "source": "user",
                        "confidence": 1.0,
                    },
                }
            }
            with (
                patch.object(
                    process,
                    "get_header_mapping",
                    return_value=learned,
                ),
                patch.object(process, "infer_headers") as infer,
            ):
                result = process.prepare_target_table(
                    str(path),
                    None,
                    {"base_url": "", "token": "", "model": ""},
                )

        self.assertEqual(result["status"], "ready")
        infer.assert_not_called()

    def test_high_confidence_unknown_headers_are_saved(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "table.xlsx"
            workbook(path, ["供货单位", "结算价"])
            suggestions = [
                {
                    "column_index": 1,
                    "header": "供货单位",
                    "target_field": "supplier_name",
                    "confidence": 0.97,
                    "reason": "供应方",
                },
                {
                    "column_index": 2,
                    "header": "结算价",
                    "target_field": "unit_price",
                    "confidence": 0.95,
                    "reason": "价格",
                },
            ]
            with (
                patch.object(
                    process,
                    "infer_headers",
                    return_value=suggestions,
                ),
                patch.object(process, "save_header_mapping") as save,
            ):
                result = process.prepare_target_table(
                    str(path),
                    None,
                    {
                        "base_url": "http://gateway",
                        "token": "",
                        "model": "m",
                    },
                )

        self.assertEqual(result["status"], "ready")
        save.assert_called_once()

    def test_low_confidence_headers_return_one_confirmation(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "table.xlsx"
            workbook(path, ["结算口径"])
            suggestions = [{
                "column_index": 1,
                "header": "结算口径",
                "target_field": "unit_price",
                "confidence": 0.61,
                "reason": "不确定",
            }]
            with patch.object(
                process,
                "infer_headers",
                return_value=suggestions,
            ):
                result = process.prepare_target_table(
                    str(path),
                    None,
                    {
                        "base_url": "http://gateway",
                        "token": "",
                        "model": "m",
                    },
                )

        self.assertEqual(result["status"], "needs_header_confirmation")
        self.assertEqual(len(result["pending_mappings"]), 1)

    def test_llm_failure_never_guesses_unknown_header(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "table.xlsx"
            workbook(path, ["结算价"])
            with patch.object(
                process,
                "infer_headers",
                return_value={
                    "status": "unavailable",
                    "reason": "invalid_response",
                },
            ):
                result = process.prepare_target_table(
                    str(path),
                    None,
                    {
                        "base_url": "http://gateway",
                        "token": "",
                        "model": "m",
                    },
                )

        self.assertEqual(result["status"], "needs_header_confirmation")
        self.assertEqual(result["mappings"], {})

    def test_all_ignored_headers_do_not_make_table_ready(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "table.xlsx"
            workbook(path, ["内部备注"])
            suggestions = [{
                "column_index": 1,
                "header": "内部备注",
                "target_field": None,
                "confidence": 0.95,
                "reason": "用户维护列",
            }]
            with patch.object(
                process,
                "infer_headers",
                return_value=suggestions,
            ):
                result = process.prepare_target_table(
                    str(path),
                    None,
                    {
                        "base_url": "http://gateway",
                        "token": "",
                        "model": "m",
                    },
                )

        self.assertEqual(result["status"], "needs_header_confirmation")

    def test_creates_target_table_from_headers(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "table.xlsx"

            created = process.create_target_table(
                str(path),
                "供货单位, 结算价, 到货量",
            )

            headers = [
                cell.value for cell in load_workbook(path).active[1]
            ]
            self.assertEqual(created, str(path))
            self.assertEqual(headers, ["供货单位", "结算价", "到货量"])

    def test_user_mapping_is_validated_and_saved(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "table.xlsx"
            workbook(path, ["供货单位", "结算价"])
            with patch.object(process, "save_header_mapping") as save:
                result = process.save_user_header_mapping(
                    str(path),
                    {"1": "supplier_name", "2": "unit_price"},
                )

            self.assertEqual(result[1]["target_field"], "supplier_name")
            save.assert_called_once()

    def test_invalid_user_mapping_is_rejected(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "table.xlsx"
            workbook(path, ["供货单位"])

            with self.assertRaisesRegex(ValueError, "不支持的目标字段"):
                process.save_user_header_mapping(
                    str(path),
                    {"1": "shell_command"},
                )

    def test_setup_table_cli_creates_table_without_images(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "table.xlsx"
            with (
                patch.object(
                    sys,
                    "argv",
                    [
                        "process.py",
                        "--setup-table",
                        str(path),
                        "--headers",
                        "单号,单价,数量",
                        "--agent-mode",
                    ],
                ),
                patch.object(
                    process,
                    "load_preferences",
                    return_value={"default_table": None},
                ),
                patch.object(process, "set_default_table") as set_default,
                patch("builtins.print") as print_mock,
                self.assertRaises(SystemExit) as exit_context,
            ):
                process.main()

            self.assertEqual(exit_context.exception.code, 0)
            self.assertTrue(path.is_file())
            set_default.assert_called_once_with(str(path))
            payload = json.loads(print_mock.call_args.args[0])
            self.assertEqual(payload["status"], "table_ready")

    def test_categorize_exports_with_table_mapping(self):
        result = valid_invoice_result("001.jpg")
        result["review_status"] = "confirmed"
        batch = {"batch_id": "b1", "results": [result]}
        mappings = {
            1: {"header": "供货单位", "target_field": "supplier_name"},
            2: {"header": "货品编号", "target_field": "fabric_code"},
        }

        with patch.object(process, "export_xlsx_append") as export:
            process._categorize(
                batch,
                Path("/tmp/templates"),
                "/tmp/table.xlsx",
                table_mappings=mappings,
            )

        export.assert_called_once()
        self.assertEqual(export.call_args.args[2], mappings)


class ResultPostProcessingTests(unittest.TestCase):
    def test_same_supplier_reuses_template_within_batch(self):
        with tempfile.TemporaryDirectory() as tmp:
            templates_dir = Path(tmp)
            first = valid_invoice_result("001.jpg")
            second = valid_invoice_result("002.jpg")
            second["data"]["delivery_note"]["note_number"] = "WT-002"

            processed = process.post_process_results(
                [second, first],
                templates_dir,
            )

        self.assertTrue(processed[0]["auto_template_created"])
        self.assertEqual(processed[1]["template_matched"], "旺泰纺织")
        self.assertEqual(
            [item["review_status"] for item in processed],
            ["confirmed", "confirmed"],
        )


if __name__ == "__main__":
    unittest.main()
