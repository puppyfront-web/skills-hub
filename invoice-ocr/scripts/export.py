#!/usr/bin/env python3
"""
Export delivery note extraction results to a user-selected Excel workbook.

Columns: 单号, 日期, 款号, 描述, 件数, 面料/辅料, 面料厂家, 面料款号, 色号/颜色,
         单价, 数量, 单位, 总金额(=J*K), 单件金额(=N/E), 用料(=K/E), 备注
Plus right-side cost breakdown: 面料, 辅料, 砂洗, 加工, 裁床, 吊牌, 包装, 合计

Supports append/update mode: preserves existing file, appends new rows,
detects price changes on same fabric from same supplier (highlights red).
"""

import argparse
import csv
import json
import sys
from dataclasses import dataclass
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Main data columns (A-P)
MAIN_HEADERS = [
    "单号",         # A - note number
    "日期",         # B - date
    "款号",         # C - style number (user fills)
    "描述",         # D - description (user fills)
    "件数",         # E - pieces (user fills)
    "面料/辅料",    # F - material type
    "面料厂家",     # G - supplier
    "面料款号",     # H - fabric code
    "色号/颜色",    # I - color code
    "单价",         # J - unit price
    "数量",         # K - quantity
    "单位",         # L - unit
    "总金额",       # M - total amount = J*K
    "单件金额",     # N - per-piece cost = M/E
    "用料",         # O - usage = K/E
    "备注",         # P - remarks
]

# Cost breakdown columns (Q-X)
COST_HEADERS = [
    "面料", "辅料", "砂洗", "加工", "裁床", "吊牌", "包装", "合计",
]

ALL_HEADERS = MAIN_HEADERS + COST_HEADERS
NUM_MAIN = len(MAIN_HEADERS)      # 16
NUM_ALL = len(ALL_HEADERS)        # 24

COL_WIDTHS = {
    "单号": 18, "日期": 12,
    "款号": 10, "描述": 18, "件数": 8, "面料/辅料": 10,
    "面料厂家": 18, "面料款号": 14, "色号/颜色": 14,
    "单价": 10, "数量": 10, "单位": 8,
    "总金额": 14, "单件金额": 12, "用料": 10, "备注": 28,
    "面料": 10, "辅料": 10, "砂洗": 10, "加工": 10,
    "裁床": 10, "吊牌": 10, "包装": 10, "合计": 12,
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
REVIEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
PRICE_ALERT_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
CLEAR_FILL = PatternFill(fill_type=None)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
NUM_FMT_MONEY = '#,##0.00'
NUM_FMT_QTY = '#,##0.##'
NUM_FMT_INT = '#,##0'


@dataclass(frozen=True)
class ExportStats:
    written_rows: int
    price_alerts: int
    skipped_duplicates: int
    output_path: str
    price_alert_details: tuple[dict, ...] = ()


def load_results(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def build_remark(item: dict, group: dict) -> str:
    """Build remark string from item metadata."""
    parts = []
    if group.get("note_number"):
        parts.append(f"单号:{group['note_number']}")
    if item.get("fabric_code_is_handwritten"):
        parts.append("手写货号")
    remark = item.get("remark") or ""
    if remark:
        parts.append(remark)
    return "; ".join(parts) if parts else ""


def iter_invoice_groups(batch: dict):
    """Yield (result_dict, group_meta) for each invoice, including errors."""
    for result in batch.get("results", []):
        if result.get("review_status") == "skipped":
            continue
        if result.get("status") == "success" and result.get("data"):
            data = result["data"]
            delivery = data.get("delivery_note", {})
            yield result, {
                "filename": result.get("filename", ""),
                "supplier_name": delivery.get("supplier_name") or "",
                "note_number": delivery.get("note_number") or "",
                "date": delivery.get("date") or "",
                "customer": delivery.get("customer") or "",
                "items": data.get("items", []),
                "total_amount": data.get("total_amount"),
                "confidence": data.get("confidence", "medium"),
                "needs_review": data.get("needs_review", []),
                "review_status": result.get("review_status", "pending"),
                "raw_text_notes": data.get("raw_text_notes") or "",
                "is_error": False,
            }
        else:
            yield result, {
                "filename": result.get("filename", ""),
                "supplier_name": "",
                "note_number": "",
                "date": "",
                "customer": "",
                "items": [],
                "total_amount": None,
                "confidence": "low",
                "needs_review": [],
                "review_status": result.get("review_status", "pending"),
                "raw_text_notes": "",
                "is_error": True,
                "error": result.get("error") or "unknown error",
            }


def _set_cell(ws, row: int, col: int, value=None, fmt=None, fill=None):
    """Write a cell with border and optional format/fill."""
    c = ws.cell(row=row, column=col, value=value)
    c.border = THIN_BORDER
    if fmt:
        c.number_format = fmt
    if fill is not None:
        c.fill = fill
    return c


def _excel_safe(value):
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _setup_header(ws):
    """Write header row with styling."""
    for col, header in enumerate(ALL_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=header)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGNMENT
        c.border = THIN_BORDER


def _setup_columns(ws):
    """Set column widths, freeze panes, auto-filter."""
    for col, header in enumerate(ALL_HEADERS, 1):
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS.get(header, 12)
    ws.freeze_panes = "A2"


def _build_price_index(ws, last_row: int) -> dict:
    """Build index of (supplier, fabric_code) -> unit_price from existing rows.

    Only indexes rows where both supplier (col G) and fabric_code (col H) exist.
    Returns dict mapping (supplier, fabric_code) -> unit_price (float or None).
    """
    prices = {}
    for row in range(2, last_row + 1):
        supplier = ws.cell(row=row, column=7).value  # G
        fabric_code = ws.cell(row=row, column=8).value  # H
        unit_price = ws.cell(row=row, column=10).value  # J
        if supplier and fabric_code:
            key = (str(supplier).strip(), str(fabric_code).strip())
            if key not in prices and unit_price is not None:
                try:
                    prices[key] = float(unit_price)
                except (TypeError, ValueError):
                    continue
    return prices


def _build_mapped_price_index(ws, last_row: int, field_to_col: dict) -> dict:
    supplier_col = field_to_col.get("supplier_name")
    fabric_col = field_to_col.get("fabric_code")
    price_col = field_to_col.get("unit_price")
    if not supplier_col or not fabric_col or not price_col:
        return {}

    prices = {}
    for row in range(2, last_row + 1):
        supplier = ws.cell(row=row, column=supplier_col).value
        fabric_code = ws.cell(row=row, column=fabric_col).value
        unit_price = ws.cell(row=row, column=price_col).value
        if supplier and fabric_code and unit_price is not None:
            key = (str(supplier).strip(), str(fabric_code).strip())
            if key not in prices:
                try:
                    prices[key] = float(unit_price)
                except (TypeError, ValueError):
                    continue
    return prices


def _write_item_row(ws, r: int, group: dict, item: dict, row_fill=None,
                    price_alert: bool = False, preserve_user_columns: bool = False):
    """Write a single item row starting at row r."""
    # A: 单号
    _set_cell(ws, r, 1, value=group.get("note_number") or "", fill=row_fill)

    # B: 日期
    _set_cell(ws, r, 2, value=group.get("date") or "", fill=row_fill)

    # C: 款号 (auto-fill if style_number present, else user fills)
    _set_cell(ws, r, 3, value=item.get("style_number") or None, fill=row_fill)

    # D-E: 描述, 件数 (user fills)
    if not preserve_user_columns:
        _set_cell(ws, r, 4, fill=row_fill)
        _set_cell(ws, r, 5, fill=row_fill)

    # F: 面料/辅料
    _set_cell(ws, r, 6, value=item.get("material_type") or "", fill=row_fill)

    # G: 面料厂家
    supplier = item.get("supplier") or group["supplier_name"] or ""
    _set_cell(ws, r, 7, value=supplier, fill=row_fill)

    # H: 面料款号
    _set_cell(ws, r, 8, value=item.get("fabric_code") or "", fill=row_fill)

    # I: 色号/颜色
    _set_cell(ws, r, 9, value=item.get("color_code") or "", fill=row_fill)

    # J: 单价 — red fill if price changed from previous entries
    j_fill = PRICE_ALERT_FILL if price_alert else row_fill
    _set_cell(ws, r, 10, value=item.get("unit_price"), fmt=NUM_FMT_MONEY, fill=j_fill)

    # K: 数量
    _set_cell(ws, r, 11, value=item.get("quantity"), fmt=NUM_FMT_QTY, fill=row_fill)

    # L: 单位
    _set_cell(ws, r, 12, value=item.get("unit") or "", fill=row_fill)

    # M: 总金额 = J*K (formula when both present, else raw value)
    unit_price = item.get("unit_price")
    quantity = item.get("quantity")
    if unit_price is not None and quantity is not None:
        _set_cell(ws, r, 13, value=f"=J{r}*K{r}", fmt=NUM_FMT_MONEY, fill=row_fill)
    else:
        _set_cell(ws, r, 13, value=item.get("total_amount"), fmt=NUM_FMT_MONEY, fill=row_fill)

    # N: 单件金额 = M/E (safe division)
    _set_cell(ws, r, 14, value=f'=IF(E{r}="","",M{r}/E{r})', fmt=NUM_FMT_MONEY, fill=row_fill)

    # O: 用料 = K/E (safe division)
    _set_cell(ws, r, 15, value=f'=IF(E{r}="","",K{r}/E{r})', fmt=NUM_FMT_QTY, fill=row_fill)

    # P: 备注
    remark = build_remark(item, group)
    _set_cell(ws, r, 16, value=remark, fill=row_fill)

    # Q-X: cost breakdown (user fills)
    if not preserve_user_columns:
        for col in range(17, NUM_ALL + 1):
            _set_cell(ws, r, col, fill=row_fill)


def _normalized_cell_value(value) -> str:
    return str(value or "").strip()


def _note_key(supplier, note_number):
    note = _normalized_cell_value(note_number)
    if not note:
        return None
    return ("note", f"{_normalized_cell_value(supplier)}|{note}")


def _item_key(supplier, fabric_code):
    return ("sf", f"{_normalized_cell_value(supplier)}|{_normalized_cell_value(fabric_code)}")


def _price_alert_for_item(price_index: dict, group: dict, item: dict) -> tuple[bool, tuple | None]:
    supplier = item.get("supplier") or group["supplier_name"] or ""
    fabric_code = item.get("fabric_code") or ""
    unit_price = item.get("unit_price")
    if not supplier or not fabric_code or unit_price is None:
        return False, None
    key = (str(supplier).strip(), str(fabric_code).strip())
    prev_price = price_index.get(key)
    if prev_price is not None and abs(float(unit_price) - prev_price) > 0.005:
        return True, key
    return False, key


def _price_alert_detail(group: dict, item: dict, previous_price: float) -> dict:
    supplier = item.get("supplier") or group["supplier_name"] or ""
    return {
        "supplier_name": str(supplier).strip(),
        "fabric_code": str(item.get("fabric_code") or "").strip(),
        "previous_price": previous_price,
        "current_price": float(item.get("unit_price")),
        "note_number": group.get("note_number") or "",
    }


def _build_note_index(ws, last_row: int) -> dict:
    """Build index of (note_number_or_filename) -> set of row numbers for dedup.

    Returns dict mapping identifier -> set of row numbers.
    Uses note_number if present, otherwise falls back to supplier+fabric_code combo.
    """
    index = {}
    for row in range(2, last_row + 1):
        note = ws.cell(row=row, column=1).value  # A: 单号
        supplier = ws.cell(row=row, column=7).value  # G: 面料厂家
        fabric_code = ws.cell(row=row, column=8).value  # H: 面料款号
        if note and str(note).strip():
            key = _note_key(supplier, note)
        else:
            # Fallback: supplier + fabric_code combo
            key = _item_key(supplier, fabric_code)
        index.setdefault(key, set()).add(row)
    return index


def _build_mapped_note_index(ws, last_row: int, field_to_col: dict) -> dict:
    note_col = field_to_col.get("note_number")
    if not note_col:
        return {}

    supplier_col = field_to_col.get("supplier_name")
    index = {}
    for row in range(2, last_row + 1):
        supplier = ws.cell(row=row, column=supplier_col).value if supplier_col else ""
        key = _note_key(supplier, ws.cell(row=row, column=note_col).value)
        if key:
            index.setdefault(key, set()).add(row)
    return index


def _is_invoice_entry_sheet(ws) -> bool:
    headers = [
        ws.cell(row=1, column=col).value
        for col in range(1, NUM_MAIN + 1)
    ]
    return headers == MAIN_HEADERS


def _prepare_invoice_sheet(wb):
    """Find or create the sheet used for invoice rows without replacing user sheets."""
    if _is_invoice_entry_sheet(wb.active):
        return wb.active, False

    for ws in wb.worksheets:
        if _is_invoice_entry_sheet(ws):
            return ws, False

    if "票据录入" in wb.sheetnames:
        ws = wb["票据录入"]
        if ws.max_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
            _setup_header(ws)
            _setup_columns(ws)
            return ws, True

    ws = wb.create_sheet("票据录入")
    _setup_header(ws)
    _setup_columns(ws)
    return ws, True


def _last_data_row(ws) -> int:
    last_row = ws.max_row
    while last_row > 1 and all(
        ws.cell(row=last_row, column=c).value is None
        for c in range(1, ws.max_column + 1)
    ):
        last_row -= 1
    return last_row


def _target_field_to_column(mappings: dict) -> dict:
    fields = {}
    for column, mapping in mappings.items():
        field = mapping.get("target_field") if isinstance(mapping, dict) else None
        if field:
            fields[field] = int(column)
    return fields


def _mapped_row_values(group: dict, item: dict) -> dict:
    unit_price = item.get("unit_price")
    quantity = item.get("quantity")
    total_amount = item.get("total_amount")
    if total_amount is None and unit_price is not None and quantity is not None:
        total_amount = unit_price * quantity

    supplier = item.get("supplier") or group.get("supplier_name") or ""
    return {
        "note_number": group.get("note_number") or "",
        "date": group.get("date") or "",
        "customer": group.get("customer") or "",
        "supplier_name": supplier,
        "style_number": item.get("style_number") or "",
        "material_type": item.get("material_type") or "",
        "material_name": item.get("material_name") or "",
        "fabric_code": item.get("fabric_code") or "",
        "color_code": item.get("color_code") or "",
        "unit_price": unit_price,
        "quantity": quantity,
        "unit": item.get("unit") or "",
        "total_amount": total_amount,
        "remark": build_remark(item, group),
    }


def _export_mapped_xlsx_append(
    batch: dict,
    output_path: str,
    mappings: dict,
    only_confirmed: bool = False,
) -> ExportStats:
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(output_path) if output_file.exists() else Workbook()
    ws = wb.active
    last_row = _last_data_row(ws)

    field_to_col = _target_field_to_column(mappings)
    price_index = _build_mapped_price_index(ws, last_row, field_to_col)
    existing_notes = _build_mapped_note_index(ws, last_row, field_to_col)

    row_num = last_row + 1
    written_rows = 0
    price_alerts = 0
    skipped_duplicates = 0
    price_alert_details = []

    for _, group in iter_invoice_groups(batch):
        if only_confirmed and group["review_status"] != "confirmed":
            continue
        if group.get("is_error") or not group.get("items"):
            continue

        note_key = _note_key(group["supplier_name"], group.get("note_number"))
        if note_key and note_key in existing_notes:
            skipped_duplicates += 1
            continue

        group_rows = []
        for item in group["items"]:
            price_alert, price_key = _price_alert_for_item(price_index, group, item)
            if price_alert:
                price_alerts += 1
                price_alert_details.append(
                    _price_alert_detail(group, item, price_index[price_key])
                )

            values = _mapped_row_values(group, item)
            for field, column in field_to_col.items():
                fill = PRICE_ALERT_FILL if field == "unit_price" and price_alert else None
                _set_cell(
                    ws,
                    row_num,
                    column,
                    value=_excel_safe(values.get(field)),
                    fill=fill,
                )

            group_rows.append(row_num)
            if price_key:
                price_index[price_key] = float(item.get("unit_price"))
            row_num += 1
            written_rows += 1

        if note_key and group_rows:
            existing_notes[note_key] = set(group_rows)

    wb.save(output_path)
    return ExportStats(
        written_rows,
        price_alerts,
        skipped_duplicates,
        output_path,
        tuple(price_alert_details),
    )


def export_xlsx_append(
    batch: dict,
    output_path: str,
    mappings: dict | bool | None = None,
    only_confirmed: bool = False,
):
    """Append new rows to existing Excel file, preserving user edits.

    If file doesn't exist, creates a new one.
    """
    if isinstance(mappings, bool):
        only_confirmed = mappings
        mappings = None
    if mappings is not None:
        return _export_mapped_xlsx_append(batch, output_path, mappings, only_confirmed)

    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    if output_file.exists():
        try:
            wb = load_workbook(output_path)
        except PermissionError:
            print(f"Error: Cannot open {output_path} (permission denied). "
                  f"Using a new file.", file=sys.stderr)
            output_path = str(Path(output_path).with_name(
                Path(output_path).stem + "_new" + Path(output_path).suffix))
            wb = Workbook()
            ws = wb.active
            ws.title = "票据录入"
            _setup_header(ws)
            _setup_columns(ws)
            last_row = 1
        else:
            ws, created_sheet = _prepare_invoice_sheet(wb)
            last_row = ws.max_row
            # Skip any trailing empty rows
            while last_row > 1 and all(
                ws.cell(row=last_row, column=c).value is None
                for c in range(1, NUM_ALL + 1)
            ):
                last_row -= 1
            if created_sheet:
                last_row = 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "票据录入"
        _setup_header(ws)
        _setup_columns(ws)
        last_row = 1

    # Build price index and note index from existing rows
    price_index = _build_price_index(ws, last_row)
    existing_notes = _build_note_index(ws, last_row)

    row_num = last_row + 1
    new_rows = 0
    price_alerts = 0
    price_alert_details = []

    for result, group in iter_invoice_groups(batch):
        if only_confirmed and group["review_status"] != "confirmed":
            continue

        items = group["items"]
        has_review = bool(group.get("needs_review"))
        row_fill = REVIEW_FILL if has_review else CLEAR_FILL

        if group.get("is_error"):
            for col in range(1, NUM_ALL + 1):
                _set_cell(ws, row_num, col, fill=ERROR_FILL)
            ws.cell(row=row_num, column=6, value="[提取失败]")
            ws.cell(row=row_num, column=7, value=group.get("filename", ""))
            ws.cell(row=row_num, column=16, value=group.get("error", ""))
            for col in range(1, NUM_ALL + 1):
                ws.cell(row=row_num, column=col).fill = ERROR_FILL
            row_num += 1
            new_rows += 1
            continue

        if not items:
            for col in range(1, NUM_ALL + 1):
                _set_cell(ws, row_num, col, fill=row_fill)
            ws.cell(row=row_num, column=1, value=group.get("note_number") or "")
            ws.cell(row=row_num, column=2, value=group.get("date") or "")
            ws.cell(row=row_num, column=7, value=group["supplier_name"] or group["filename"])
            ws.cell(row=row_num, column=16, value=f"[无明细] {group.get('raw_text_notes', '')}")
            row_num += 1
            new_rows += 1
            continue

        # Check if this note was already imported
        note_number = group.get("note_number") or ""
        note_key = _note_key(group["supplier_name"], note_number)
        existing_note_rows = sorted(existing_notes.get(note_key, [])) if note_key else []
        if existing_note_rows:
            group_rows = []
            for idx, item in enumerate(items):
                price_alert, price_key = _price_alert_for_item(price_index, group, item)
                if price_alert:
                    price_alerts += 1
                    price_alert_details.append(
                        _price_alert_detail(group, item, price_index[price_key])
                    )

                if idx < len(existing_note_rows):
                    target_row = existing_note_rows[idx]
                    _write_item_row(
                        ws, target_row, group, item, row_fill, price_alert,
                        preserve_user_columns=True,
                    )
                    group_rows.append(target_row)
                else:
                    _write_item_row(ws, row_num, group, item, row_fill, price_alert)
                    group_rows.append(row_num)
                    row_num += 1
                    new_rows += 1

                if price_key:
                    price_index[price_key] = float(item.get("unit_price"))

            existing_notes[note_key] = set(group_rows)
            continue

        wrote_group = False
        group_rows = []
        for item in items:
            item_key = None
            if not note_key:
                supplier = str(item.get("supplier") or group["supplier_name"] or "").strip()
                fabric_code = str(item.get("fabric_code") or "").strip()
                item_key = _item_key(supplier, fabric_code)
                if item_key in existing_notes:
                    target_row = min(existing_notes[item_key])
                    price_alert, price_key = _price_alert_for_item(price_index, group, item)
                    if price_alert:
                        price_alerts += 1
                        price_alert_details.append(
                            _price_alert_detail(group, item, price_index[price_key])
                        )
                    _write_item_row(
                        ws, target_row, group, item, row_fill, price_alert,
                        preserve_user_columns=True,
                    )
                    if price_key:
                        price_index[price_key] = float(item.get("unit_price"))
                    continue

            # Check price change for same supplier + fabric_code
            price_alert, price_key = _price_alert_for_item(price_index, group, item)
            if price_alert:
                price_alerts += 1
                price_alert_details.append(
                    _price_alert_detail(group, item, price_index[price_key])
                )

            _write_item_row(ws, row_num, group, item, row_fill, price_alert)
            wrote_group = True
            group_rows.append(row_num)
            if item_key:
                existing_notes.setdefault(item_key, set()).add(row_num)
            if price_key:
                price_index[price_key] = float(item.get("unit_price"))
            row_num += 1
            new_rows += 1

        if note_key and wrote_group:
            existing_notes.setdefault(note_key, set()).update(group_rows)

    # Update column widths (only if new file)
    if last_row == 1:
        _setup_columns(ws)

    # Update auto-filter
    final_row = row_num - 1
    if final_row >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(NUM_ALL)}{final_row}"

    try:
        wb.save(output_path)
    except PermissionError:
        alt = str(Path(output_path).with_name(
            Path(output_path).stem + "_new" + Path(output_path).suffix))
        print(f"Warning: Cannot write to {output_path}, saving to {alt}", file=sys.stderr)
        wb.save(alt)
        output_path = alt
    return ExportStats(new_rows, price_alerts, 0, output_path, tuple(price_alert_details))


def export_xlsx_full(batch: dict, output_path: str, only_confirmed: bool = False):
    """Full regeneration — creates new file from scratch."""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "票据录入"

    _setup_header(ws)

    row_num = 2

    for result, group in iter_invoice_groups(batch):
        if only_confirmed and group["review_status"] != "confirmed":
            continue

        items = group["items"]
        has_review = bool(group.get("needs_review"))
        row_fill = REVIEW_FILL if has_review else CLEAR_FILL

        if group.get("is_error"):
            for col in range(1, NUM_ALL + 1):
                _set_cell(ws, row_num, col, fill=ERROR_FILL)
            ws.cell(row=row_num, column=6, value="[提取失败]")
            ws.cell(row=row_num, column=7, value=group.get("filename", ""))
            ws.cell(row=row_num, column=16, value=group.get("error", ""))
            for col in range(1, NUM_ALL + 1):
                ws.cell(row=row_num, column=col).fill = ERROR_FILL
            row_num += 1
            continue

        if not items:
            for col in range(1, NUM_ALL + 1):
                _set_cell(ws, row_num, col, fill=row_fill)
            ws.cell(row=row_num, column=1, value=group.get("note_number") or "")
            ws.cell(row=row_num, column=2, value=group.get("date") or "")
            ws.cell(row=row_num, column=7, value=group["supplier_name"] or group["filename"])
            ws.cell(row=row_num, column=16, value=f"[无明细] {group.get('raw_text_notes', '')}")
            row_num += 1
            continue

        for item in items:
            _write_item_row(ws, row_num, group, item, row_fill)
            row_num += 1

    _setup_columns(ws)

    last_row = row_num - 1
    if last_row >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(NUM_ALL)}{last_row}"

    wb.save(output_path)
    return last_row - 1, 0


def export_csv(batch: dict, output_path: str, only_confirmed: bool = False):
    """Export to CSV (formulas become computed values)."""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(ALL_HEADERS)

        for result, group in iter_invoice_groups(batch):
            if only_confirmed and group["review_status"] != "confirmed":
                continue

            if group.get("is_error"):
                row = [""] * NUM_ALL
                row[5] = "[提取失败]"
                row[6] = group.get("filename", "")
                row[15] = group.get("error", "")
                writer.writerow(row)
                continue

            for item in group["items"]:
                unit_price = item.get("unit_price")
                quantity = item.get("quantity")
                total = (
                    unit_price * quantity
                    if unit_price is not None and quantity is not None
                    else item.get("total_amount")
                )

                writer.writerow([
                    group.get("note_number", ""),
                    group.get("date", ""),
                    item.get("style_number", ""),  # C: 款号
                    "", "",  # D-E: 描述, 件数
                    item.get("material_type", ""),
                    item.get("supplier") or group["supplier_name"],
                    item.get("fabric_code", ""),
                    item.get("color_code", ""),
                    unit_price if unit_price is not None else "",
                    quantity if quantity is not None else "",
                    item.get("unit", ""),
                    total if total is not None else "",
                    "", "",  # N, O (need pieces)
                    build_remark(item, group),
                ] + [""] * 8)

            if not group["items"]:
                row = [""] * NUM_ALL
                row[0] = group.get("note_number", "")
                row[1] = group.get("date", "")
                row[6] = group["supplier_name"] or group["filename"]
                row[15] = f"[无明细] {group.get('raw_text_notes', '')}"
                writer.writerow(row)


def main():
    parser = argparse.ArgumentParser(description="Export delivery note results to Excel/CSV")
    parser.add_argument("input_json", help="Results JSON file from extract.py")
    parser.add_argument("--output", "-o", help="Output file path (overrides preferences)")
    parser.add_argument("--format", "-f", choices=["xlsx", "csv"], default="xlsx")
    parser.add_argument("--mode", "-m", choices=["append", "full"], default="append",
                        help="Export mode: append (default, preserves existing file) or full (regenerate)")
    parser.add_argument("--only-confirmed", action="store_true",
                        help="Only export confirmed/reviewed invoices")
    args = parser.parse_args()

    batch = load_results(args.input_json)

    # Resolve output path: explicit arg > supplier preference > default preference > same dir
    if args.output:
        output_path = args.output
    else:
        sys.path.insert(0, str(Path(__file__).parent))
        from templates import get_table_for_supplier
        # Try to find a table preference from any supplier in the batch
        output_path = None
        for result in batch.get("results", []):
            if result.get("status") == "success" and result.get("data"):
                supplier = result["data"].get("delivery_note", {}).get("supplier_name", "")
                if supplier:
                    pref = get_table_for_supplier(supplier)
                    if pref:
                        output_path = pref
                        break
        if not output_path:
            print(
                "Error: No output table specified. Pass --output with the target Excel/CSV file, "
                "or set a saved table preference first.",
                file=sys.stderr,
            )
            sys.exit(1)

    if args.format == "csv":
        export_csv(batch, output_path, args.only_confirmed)
        rows = sum(1 for _, g in iter_invoice_groups(batch)
                   if not args.only_confirmed or g["review_status"] == "confirmed")
        print(f"Exported {rows} rows to {output_path}")
    elif args.mode == "append":
        stats = export_xlsx_append(batch, output_path, args.only_confirmed)
        confirmed = batch.get("summary", {}).get("confirmed", 0)
        pending = batch.get("summary", {}).get("pending_review", 0)
        print(f"Appended {stats.written_rows} rows to {stats.output_path}")
        if stats.price_alerts:
            print(f"  {stats.price_alerts} price change(s) detected (highlighted red)")
        print(f"  {confirmed} confirmed, {pending} pending review")
    else:
        rows, _ = export_xlsx_full(batch, output_path, args.only_confirmed)
        confirmed = batch.get("summary", {}).get("confirmed", 0)
        pending = batch.get("summary", {}).get("pending_review", 0)
        print(f"Exported {rows} rows to {output_path}")
        print(f"  {confirmed} confirmed, {pending} pending review")


if __name__ == "__main__":
    main()
