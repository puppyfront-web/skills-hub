#!/usr/bin/env python3
"""
Unified pipeline for invoice OCR: extract → post-process → review → export.

One-command entry point for the full workflow:
  python3 process.py ~/photos/ -o ~/采购入库4月.xlsx           # process + review
  python3 process.py ~/photos/ -o ~/采购入库4月.xlsx --finalize  # final export
"""

import argparse
import json
import os
import sys
import threading
from pathlib import Path

# Import from sibling modules
sys.path.insert(0, str(Path(__file__).parent))
from extract import (
    check_gateway, check_model_vision_support, collect_images, extract_single,
    load_existing_results, merge_results, EXTRACT_PROMPT,
    DEFAULT_GATEWAY_HOST, DEFAULT_GATEWAY_PORT, MAX_RETRIES,
)
from templates import (
    TEMPLATES_DIR, load_index, match_template, post_process_extraction,
    format_extraction_review, parse_natural_correction,
    save_from_results, get_table_for_supplier, set_supplier_table, set_default_table,
    load_preferences, REVIEW_COLUMNS, FIELD_LABELS,
    build_learning_entries, get_header_mapping, save_header_mapping,
    auto_confirm_and_learn,
)
from export import export_xlsx_append, iter_invoice_groups
from header_mapper import infer_headers
from table_schema import (
    ALLOWED_FIELDS,
    exact_mappings,
    header_signature,
    normalize_header,
    resolve_header_mapping,
)

ASSISTANT_SUMMARY_PREFIX = "ASSISTANT_SUMMARY_JSON:"
ASSISTANT_PROGRESS_PREFIX = "ASSISTANT_PROGRESS_JSON:"


def read_headers(path: str) -> tuple[str, list[str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True)
    sheet = workbook.active
    headers = [
        normalize_header(sheet.cell(1, column).value)
        for column in range(1, sheet.max_column + 1)
    ]
    workbook.close()
    return sheet.title, headers


def create_target_table(path: str, header_text: str) -> str:
    from openpyxl import Workbook

    headers = [
        normalize_header(value)
        for value in header_text.split(",")
        if normalize_header(value)
    ]
    if not headers:
        raise ValueError("请提供至少一个表头字段")

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.append(headers)
    workbook.save(target)
    return str(target)


def save_user_header_mapping(
    path: str,
    user_mapping: dict[str, str | None],
) -> dict[int, dict]:
    _, headers = read_headers(path)
    mappings = {}
    occupied = set()

    for raw_column, target_field in user_mapping.items():
        try:
            column_index = int(raw_column)
        except (TypeError, ValueError) as error:
            raise ValueError(f"无效列序号: {raw_column}") from error
        if column_index < 1 or column_index > len(headers):
            raise ValueError(f"列序号超出范围: {column_index}")
        if target_field is not None and target_field not in ALLOWED_FIELDS:
            raise ValueError(f"不支持的目标字段: {target_field}")
        if target_field is not None and target_field in occupied:
            raise ValueError(f"目标字段重复: {target_field}")
        if target_field is not None:
            occupied.add(target_field)
        mappings[column_index] = {
            "column_index": column_index,
            "header": headers[column_index - 1],
            "target_field": target_field,
            "source": "user",
            "confidence": 1.0,
            "reason": "用户确认",
        }

    signature = header_signature(headers)
    save_header_mapping(signature, headers, mappings)
    return mappings


def prepare_target_table(
    explicit: str | None,
    saved: str | None,
    gateway: dict,
) -> dict:
    path = explicit or saved
    if not path or not Path(path).is_file():
        return {"status": "waiting_for_table", "path": None}

    sheet_name, headers = read_headers(path)
    signature = header_signature(headers)
    record = get_header_mapping(signature)
    learned = record.get("mappings", {}) if record else {}
    exact = exact_mappings(headers)
    unresolved = [
        {"column_index": index, "header": header}
        for index, header in enumerate(headers, 1)
        if str(index) not in learned and index not in exact
    ]

    suggestions = []
    llm_unavailable = False
    if unresolved:
        suggestions = infer_headers(
            gateway.get("base_url", ""),
            gateway.get("token", ""),
            gateway.get("model", "openclaw/default"),
            unresolved,
            Path(path).name,
        )
        if isinstance(suggestions, dict):
            llm_unavailable = True
            suggestions = []

    mappings, pending, _ = resolve_header_mapping(
        headers,
        learned,
        suggestions,
    )
    if llm_unavailable:
        pending.extend({
            **item,
            "target_field": None,
            "confidence": 0.0,
            "reason_code": "llm_unavailable",
        } for item in unresolved)

    result = {
        "path": path,
        "sheet_name": sheet_name,
        "signature": signature,
        "mappings": mappings,
        "pending_mappings": pending,
    }
    has_writable_field = any(
        mapping.get("target_field")
        for mapping in mappings.values()
    )
    if not has_writable_field:
        return {**result, "status": "needs_header_confirmation"}

    save_header_mapping(signature, headers, mappings)
    status = "ready_with_pending_mapping" if pending else "ready"
    return {**result, "status": status}


def post_process_results(
    results: list[dict],
    templates_dir: Path,
) -> list[dict]:
    processed = []
    for result in sorted(results, key=lambda item: item["filename"]):
        if result.get("status") == "success":
            result = post_process_extraction(result, templates_dir)
            result = auto_confirm_and_learn(result, templates_dir)
        processed.append(result)
    return processed


def process_batch(images_dir: str, output: str,
                  gateway_host: str, gateway_port: int, token: str,
                  model: str, templates_dir: Path,
                  parallel: int = 3, retries: int = 2,
                  append: bool = False,
                  review_in_excel: bool = False,
                  assistant_progress: bool = False,
                  table_mappings: dict | None = None,
                  pending_mappings: list[dict] | None = None) -> dict:
    """Full pipeline: extract → post-process → categorize results.

    Returns dict with: confirmed, pending, results, output_path.
    """
    base_url = f"http://{gateway_host}:{gateway_port}"
    check_gateway(base_url, token)

    # Build supplier context
    from templates import build_supplier_context
    prompt_extra = build_supplier_context(templates_dir)

    # Collect images
    all_images = collect_images(images_dir)

    # Append mode: skip already-processed
    results_json_path = str(Path(images_dir) / "results.json")
    existing_filenames = set()
    if append:
        existing = load_existing_results(results_json_path)
        if existing:
            existing_filenames = {r["filename"] for r in existing.get("results", [])}

    images = [img for img in all_images
              if os.path.basename(img) not in existing_filenames]

    if not images:
        print("No new images to process.")
        # Still load existing results for export
        if Path(results_json_path).is_file():
            with open(results_json_path, "r", encoding="utf-8") as f:
                batch = json.load(f)
            return _categorize(
                batch,
                templates_dir,
                output,
                export_pending=review_in_excel,
                table_mappings=table_mappings,
                pending_mappings=pending_mappings,
            )
        return {
            "confirmed": [],
            "pending": [],
            "batch": None,
            "output_path": output,
            "confirmed_suppliers": [],
            "exported_pending": False,
            "nothing_to_process": True,
        }

    # Extract
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from datetime import datetime

    print(f"Processing {len(images)} delivery notes...")
    progress_lock = threading.Lock()

    def emit_progress(status: str, message: str, **extra):
        if not assistant_progress:
            return
        with progress_lock:
            print_assistant_progress({
                "status": status,
                "user_message": message,
                "counts": extra.pop("counts", {}),
                **extra,
            })

    def extract_progress(filename: str, stage: str, message: str):
        emit_progress(stage, message, filename=filename)

    emit_progress(
        "started",
        f"收到，我先帮你批量识别这 {len(images)} 张票据，等下把需要确认的一起发你。",
        counts={"total": len(images), "done": 0},
    )
    results = []
    with ThreadPoolExecutor(max_workers=parallel) as executor:
        futures = {executor.submit(extract_single, base_url, token,
                                   img, model, retries, prompt_extra,
                                   extract_progress if assistant_progress else None): img
                   for img in images}
        for future in as_completed(futures):
            img = futures[future]
            try:
                result = future.result()
                results.append(result)
            except Exception as e:
                results.append({
                    "filename": os.path.basename(img), "status": "error",
                    "error": str(e), "data": None, "review_status": "pending",
                })
            done = len(results)
            emit_progress(
                "image_done",
                f"已处理 {done}/{len(images)} 张票据。",
                counts={"total": len(images), "done": done},
                filename=os.path.basename(img),
            )

    results = post_process_results(results, templates_dir)
    batch_id = Path(images_dir).name

    # Merge or create batch
    if append and existing_filenames:
        existing = load_existing_results(results_json_path) or {}
        batch = merge_results(existing, results, batch_id)
    else:
        confirmed_count = sum(1 for r in results if r.get("review_status") == "confirmed")
        pending_count = sum(1 for r in results if r.get("review_status") == "pending")
        batch = {
            "batch_id": batch_id,
            "created_at": datetime.now().isoformat(),
            "total_images": len(results),
            "results": results,
            "summary": {
                "total_invoices": len(results),
                "confirmed": confirmed_count,
                "pending_review": pending_count,
            },
        }

    # Save results JSON
    with open(results_json_path, "w", encoding="utf-8") as f:
        json.dump(batch, f, ensure_ascii=False, indent=2)

    emit_progress(
        "writing_table",
        "识别结果已整理好，正在写入表格并生成复核提示。",
        counts={"total": len(results), "done": len(results)},
    )
    return _categorize(
        batch,
        templates_dir,
        output,
        export_pending=review_in_excel,
        table_mappings=table_mappings,
        pending_mappings=pending_mappings,
    )


def _categorize(batch: dict, templates_dir: Path, output: str,
                export_pending: bool = False,
                table_mappings: dict | None = None,
                pending_mappings: list[dict] | None = None) -> dict:
    """Split results into confirmed/pending and export according to mode."""
    result_view = build_result_view(batch, output, export_pending)
    result_view["pending_mappings"] = pending_mappings or []
    export_results = result_view["confirmed"]
    if export_pending:
        export_results = [
            result for result in batch.get("results", [])
            if result.get("status") == "success"
        ]

    if export_results:
        export_batch = {
            "batch_id": batch.get("batch_id", ""),
            "results": export_results,
            "summary": {
                "total_invoices": len(export_results),
                "confirmed": len(result_view["confirmed"]),
                "pending_review": len(result_view["pending"]),
            },
        }
        export_xlsx_append(export_batch, output, table_mappings)

    return result_view


def build_result_view(batch: dict, output: str, exported_pending: bool = False) -> dict:
    """Split batch results without writing files."""
    confirmed_results = []
    pending_results = []
    confirmed_suppliers = []
    learned = []

    for result in batch.get("results", []):
        if result.get("review_status") == "skipped":
            continue
        for item in result.get("learned", []):
            if item and item not in learned:
                learned.append(item)
        if result.get("auto_confirmed"):
            confirmed_results.append(result)
            supplier = result.get("data", {}).get("delivery_note", {}).get("supplier_name", "")
            confirmed_suppliers.append(supplier)
        elif result.get("review_status") == "confirmed":
            confirmed_results.append(result)
            supplier = result.get("data", {}).get("delivery_note", {}).get("supplier_name", "")
            confirmed_suppliers.append(supplier)
        else:
            pending_results.append(result)

    return {
        "confirmed": confirmed_results,
        "pending": pending_results,
        "batch": batch,
        "output_path": output,
        "confirmed_suppliers": confirmed_suppliers,
        "exported_pending": exported_pending,
        "learned": learned,
    }


def _result_supplier(result: dict) -> str:
    return result.get("data", {}).get("delivery_note", {}).get("supplier_name", "") or "未知供应商"


def _invoice_amount(result: dict) -> float:
    data = result.get("data") or {}
    amount = data.get("total_amount")
    try:
        return float(amount or 0)
    except (TypeError, ValueError):
        return 0.0


def _review_details(pending_success: list[dict], pending_mappings: list[dict]) -> str:
    details = []
    for item in pending_success[:3]:
        supplier = _result_supplier(item)
        issues = item.get("data", {}).get("needs_review", []) or ["需要确认"]
        details.append(f"{supplier}：{'、'.join(str(issue) for issue in issues[:2])}")

    header_names = [
        str(mapping.get("header"))
        for mapping in pending_mappings[:3]
        if mapping.get("header")
    ]
    if header_names:
        details.append(f"表头待确认：{'、'.join(header_names)}")
    return "；".join(details)


def build_assistant_summary(result: dict) -> dict:
    """Build a channel-friendly summary for OpenClaw/Feishu assistants.

    This is intentionally user-facing: channel handlers can send user_message
    directly without exposing paths, commands, or JSON internals.
    """
    confirmed = result.get("confirmed", [])
    pending = result.get("pending", [])
    output_path = result.get("output_path") or ""
    table_name = Path(output_path).name if output_path else ""
    failed = [r for r in pending if r.get("status") == "error"]
    pending_success = [r for r in pending if r.get("status") != "error"]
    total = len(confirmed) + len(pending)
    amount = sum(_invoice_amount(r) for r in confirmed + pending_success)
    pending_suppliers = sorted({_result_supplier(r) for r in pending_success})
    learned = result.get("learned", [])
    pending_mappings = result.get("pending_mappings", [])

    if result.get("needs_clarification"):
        status = "needs_clarification"
        clarification = result["needs_clarification"]
        message = clarification.get("message") or "这句话我还不知道要改哪一张，请说明供应商或单号。"
        suggested = clarification.get("candidate_suppliers") or ["说明供应商", "说明单号"]
    elif result.get("nothing_to_process"):
        status = "no_input"
        message = "没有看到新的票据照片。请把要录入的图片发给我，或确认一下文件夹里有没有新图片。"
        suggested = ["发送票据照片", "换一个文件夹"]
    elif pending_success:
        status = "needs_review"
        if result.get("exported_pending"):
            message = (
                f"已先写入表格，共 {total} 张票据，其中 {len(pending_success)} 张需要你看一下。"
                "你可以直接回复“确认”“全部确认”，或像“这家没有色号”“品名就是面料款号”这样纠正我。"
            )
        else:
            message = (
                f"识别完成，共 {total} 张票据，其中 {len(pending_success)} 张需要你确认一下。"
                "你可以直接回复“确认”“全部确认”，或告诉我哪一列不对。"
            )
        details = _review_details(pending_success, pending_mappings)
        if details:
            message = f"{message} 待确认：{details}。"
        suggested = ["确认", "全部确认", "这家没有色号", "品名就是面料款号"]
    elif failed and confirmed:
        status = "partial_success"
        message = (
            f"已录入 {len(confirmed)} 张票据，另有 {len(failed)} 张识别失败。"
            "请重新拍清楚一点再发给我，我会接着处理。"
        )
        suggested = ["重新发送照片", "先确认已录入的"]
    elif failed:
        status = "failed"
        message = "这批照片暂时没有识别成功。请重新拍清楚一点，尽量正着拍、不要裁掉金额行。"
        suggested = ["重新发送照片"]
    else:
        status = "completed"
        amount_text = f"，总金额约 ¥{amount:,.2f}" if amount else ""
        table_text = f"，已保存到 {table_name}" if table_name else ""
        message = f"已全部录入完成，共 {len(confirmed)} 张票据{amount_text}{table_text}。"
        suggested = ["继续发送下一批", "更换表格"]

    if learned:
        message = f"{message} 我记住了：{'；'.join(learned)}。"

    return {
        "status": status,
        "user_message": message,
        "counts": {
            "total": total,
            "confirmed": len(confirmed),
            "pending": len(pending_success),
            "failed": len(failed),
        },
        "table_name": table_name,
        "amount": round(amount, 2),
        "pending_suppliers": pending_suppliers,
        "suggested_replies": suggested,
        "learned": learned,
        "pending_mappings": pending_mappings,
        "needs_clarification": result.get("needs_clarification"),
    }


def build_intro_summary(has_saved_table: bool = False, template_count: int = 0) -> dict:
    table_text = "我会继续用你上次的目标 Excel。" if has_saved_table else "你把目标 Excel 发给我，或告诉我表名。"
    memory_text = (
        f"我已经记着 {template_count} 家供应商的习惯。"
        if template_count else
        "第一次遇到新样子的单子，我会请你确认几处。"
    )
    return {
        "status": "intro",
        "user_message": (
            "我是票据录入助手，可以帮你把送货单、码单、加工单照片录进表格。"
            "你只要发清楚的票据照片，多张一起也行。"
            f"{table_text}"
            "不对的地方直接说“确认”“全部确认”“第二行数量是300”“这家没有色号”。"
            f"{memory_text}你纠正过的习惯我会记下来，越用越懂。"
        ),
        "counts": {"templates": template_count},
        "suggested_replies": ["发送票据照片", "发送目标 Excel", "教我怎么纠错"],
    }


def build_need_table_summary() -> dict:
    return {
        "status": "need_table",
        "user_message": "这批要录到哪张表？请直接把目标表发给我，或告诉我表名/路径。",
        "counts": {"total": 0, "confirmed": 0, "pending": 0, "failed": 0},
        "table_name": "",
        "amount": 0,
        "pending_suppliers": [],
        "suggested_replies": ["发送目标表", "告诉我表名或路径"],
    }


def build_self_check_summary(check: dict) -> dict:
    if check.get("vision_supported"):
        status = "ready"
        message = (
            "安装检查通过：当前模型能识别图片。"
            "可以让用户发送票据照片和目标 Excel，处理时我会持续提示进度。"
        )
        suggested = ["开始发送票据照片", "发送目标 Excel"]
    else:
        status = "vision_unavailable"
        reason = check.get("reason") or "unknown"
        message = (
            "安装检查未通过：当前模型看不到图片。"
            "建议先切换到支持图片的模型后再交给票据人员使用；否则只能尝试本地 OCR 兜底，速度会慢且不稳定。"
        )
        suggested = ["切换支持图片的模型", "重新检查"]
        if reason == "gateway_error":
            message = "安装检查未通过：OpenClaw 网关暂时不可用。请先启动网关，再重新检查。"
            suggested = ["启动网关", "重新检查"]

    return {
        "status": status,
        "user_message": message,
        "vision_supported": bool(check.get("vision_supported")),
        "reason": check.get("reason"),
        "suggested_replies": suggested,
    }


def print_assistant_summary(summary: dict, *, clean: bool = False):
    """Print summary. When clean=True, output raw JSON without prefix (for --agent-mode)."""
    raw = json.dumps(summary, ensure_ascii=False, separators=(',', ':'))
    if clean:
        print(raw)
    else:
        print(f"{ASSISTANT_SUMMARY_PREFIX}{raw}")


def print_assistant_progress(progress: dict):
    print(f"{ASSISTANT_PROGRESS_PREFIX}{json.dumps(progress, ensure_ascii=False, separators=(',', ':'))}", flush=True)


def _run_interactive_loop(batch: dict, output: str, results_json: str,
                          templates_dir: Path) -> dict:
    """Interactive terminal mode: walk through pending items one by one."""
    pending = [
        r for r in batch.get("results", [])
        if r.get("status") == "success" and r.get("review_status") != "confirmed"
        and r.get("review_status") != "skipped"
    ]
    if not pending:
        return batch

    print()
    print("=" * 60)
    print("📄 交互模式：请逐项确认")
    print("=" * 60)

    for i, result in enumerate(pending):
        supplier = result.get("data", {}).get("delivery_note", {}).get("supplier_name", "未知供应商")
        filename = result.get("filename", "")
        print()
        print(f"--- [{i+1}/{len(pending)}] {supplier} ({filename}) ---")

        # Show missing fields
        data = result.get("data", {})
        items = data.get("items", [])
        is_new = not (supplier and match_template(supplier, templates_dir))

        if is_new:
            print(f"  🆕 新供应商: {supplier}")

        missing = []
        for item in items:
            for f in ["fabric_code", "color_code", "unit_price", "quantity"]:
                if item.get(f) is None or item.get(f) == "":
                    label = {"fabric_code": "面料款号", "color_code": "色号/颜色",
                             "unit_price": "单价", "quantity": "数量"}[f]
                    if label not in missing:
                        missing.append(label)

        if missing:
            print(f"  ⚠️  缺失: {', '.join(missing)}")
        print(f"  📊 {len(items)} 行明细，金额 ¥{data.get('total_amount', '?') or '?'}")

        # Interactive confirmation
        print()
        if is_new:
            print("  操作:")
            print("    y  / 回车  → 确认并记住这家")
            print("    n        → 跳过（不导入）")
            print("    e        → 编辑字段映射")
            ans = input(f"  ✅ 确认 {supplier}？[Y/n/e]: ").strip().lower()
        else:
            print("  操作:")
            print("    y  / 回车  → 确认")
            print("    n        → 跳过")
            print("    字段名=值  → 纠正，如: 色号=无色号")
            ans = input(f"  ✅ 确认 {supplier}？[Y/n/字段=值]: ").strip().lower()

        if ans == "" or ans == "y" or ans == "yes":
            result["review_status"] = "confirmed"
            result["user_confirmed"] = True
            with open(results_json, "w", encoding="utf-8") as f:
                json.dump(batch, f, ensure_ascii=False, indent=2)
            save_from_results(results_json, templates_dir, supplier)
            print(f"  ✅ {supplier} 已确认")

        elif ans == "n" or ans == "no":
            result["review_status"] = "skipped"
            with open(results_json, "w", encoding="utf-8") as f:
                json.dump(batch, f, ensure_ascii=False, indent=2)
            print(f"  ⏭️  {supplier} 已跳过")

        elif "=" in ans:
            field, value = ans.split("=", 1)
            field = field.strip()
            value = value.strip()
            # Map Chinese field labels to internal keys
            field_map = {
                "色号": "color_code", "颜色": "color_code", "色号/颜色": "color_code",
                "面料款号": "fabric_code", "货号": "fabric_code",
                "单价": "unit_price", "数量": "quantity",
                "品名": "material_name", "面料/辅料": "material_type",
            }
            internal_field = field_map.get(field, field)

            for item in items:
                if internal_field in item:
                    item[internal_field] = value

            # For special corrections like "无色号"
            if value in ("无色号", "无", "null", ""):
                pass  # Already set above

            result["review_status"] = "confirmed"
            result["user_confirmed"] = True
            with open(results_json, "w", encoding="utf-8") as f:
                json.dump(batch, f, ensure_ascii=False, indent=2)
            save_from_results(results_json, templates_dir, supplier)
            print(f"  ✅ {supplier} 已确认并记录纠正: {field} → {value}")

        elif ans == "e":
            print("  编辑模式：输入字段映射")
            print("  例: 品名=面料款号  表示『品名列是面料款号』")
            print("  例: 无色号         表示『这家没有色号』")
            mapping = input("  > ").strip()
            if mapping:
                if "=" in mapping:
                    k, v = mapping.split("=", 1)
                    corrections = [f"{k.strip()}={v.strip()}"]
                else:
                    corrections = [mapping]
                changed = apply_finalize_actions(batch, results_json, templates_dir,
                                                 corrections, supplier)
                print(f"  ✅ 已应用 {changed} 条纠正")

    print()
    print("=" * 60)
    confirmed = sum(1 for r in batch.get("results", []) if r.get("review_status") == "confirmed")
    skipped = sum(1 for r in batch.get("results", []) if r.get("review_status") == "skipped")
    print(f"📊 汇总: {confirmed} 确认, {skipped} 跳过")
    print("=" * 60)

    return batch


def format_batch_review(pending_results: list, templates_dir: Path) -> str:
    """Generate batch review summary for all pending invoices."""
    if not pending_results:
        return ""

    lines = []
    new_suppliers = []
    problem_suppliers = []

    for result in pending_results:
        if result.get("status") == "error":
            continue
        supplier = result.get("data", {}).get("delivery_note", {}).get("supplier_name", "")
        tmpl = match_template(supplier, templates_dir) if supplier else None
        if tmpl:
            problem_suppliers.append((result, tmpl))
        else:
            new_suppliers.append(result)

    if new_suppliers:
        lines.append("### 新供应商 (首次提取)")
        lines.append("")
        for result in new_suppliers:
            lines.append(format_extraction_review(result, None))
            lines.append("---")
            lines.append("")

    if problem_suppliers:
        lines.append("### 已有模版但有问题")
        lines.append("")
        for result, tmpl in problem_suppliers:
            lines.append(format_extraction_review(result, tmpl))
            lines.append("---")
            lines.append("")

    errors = [r for r in pending_results if r.get("status") == "error"]
    if errors:
        lines.append("### 提取失败")
        for r in errors:
            lines.append(f"- {r.get('filename', '')}: {r.get('error', 'unknown')}")
        lines.append("")

    return "\n".join(lines)


def print_summary(result: dict, templates_dir: Path = TEMPLATES_DIR):
    """Print human-readable processing summary."""
    confirmed = result["confirmed"]
    pending = result["pending"]
    total = len(confirmed) + len(pending)

    print(f"\n{'='*50}")
    print(f"处理完成: {total} 张票据")

    if confirmed:
        from collections import Counter
        supplier_counts = Counter(result["confirmed_suppliers"])
        supplier_str = ", ".join(f"{s}×{c}" if c > 1 else s for s, c in supplier_counts.items())
        print(f"  自动录入: {len(confirmed)} 张 ({supplier_str})")

    if pending:
        print(f"  需确认: {len(pending)} 张")
        if result.get("exported_pending"):
            print(f"  已写入Excel待复核: {result['output_path']}")
        review = format_batch_review(pending, templates_dir)
        if review:
            print(f"\n{review}")
            print("请逐一确认。可以回复：")
            print("  - '确认' → 保存规则并保留当前录入")
            print("  - '品名就是面料款号' → 记录字段映射")
            print("  - '无色号' → 标记无此字段")
    else:
        print(f"\n全部自动录入完成！Excel已保存到: {result['output_path']}")


def _result_matches_supplier(result: dict, supplier: str | None) -> bool:
    if not supplier:
        return True
    actual = _result_supplier(result)
    return actual.lower().strip() == supplier.lower().strip()


def _numeric_or_none(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _refresh_item_total(item: dict):
    unit_price = _numeric_or_none(item.get("unit_price"))
    quantity = _numeric_or_none(item.get("quantity"))
    if unit_price is not None and quantity is not None:
        total = unit_price * quantity
        item["total_amount"] = int(total) if float(total).is_integer() else total


def _refresh_delivery_total(data: dict):
    totals = []
    for item in data.get("items", []):
        total = _numeric_or_none(item.get("total_amount"))
        if total is None:
            return
        totals.append(total)
    if totals:
        total = sum(totals)
        data["total_amount"] = int(total) if float(total).is_integer() else total


def _apply_row_update(batch: dict, correction: dict, target_supplier: str | None) -> int:
    changed = 0
    row_number = correction.get("row_number")
    field = correction.get("field")
    value = correction.get("value")
    if not row_number or not field:
        return 0
    for result in batch.get("results", []):
        if result.get("status") != "success" or not result.get("data"):
            continue
        if not _result_matches_supplier(result, target_supplier or correction.get("supplier")):
            continue
        items = result["data"].get("items", [])
        idx = int(row_number) - 1
        if idx < 0 or idx >= len(items):
            continue
        items[idx][field] = value
        _refresh_item_total(items[idx])
        _refresh_delivery_total(result["data"])
        changed += 1
    return changed


def _mark_matching_results(batch: dict, target_supplier: str | None,
                           status: str, confirm_success_only: bool = True) -> int:
    changed = 0
    for result in batch.get("results", []):
        if confirm_success_only and result.get("status") != "success":
            continue
        if result.get("status") == "success" and not result.get("data"):
            continue
        if not _result_matches_supplier(result, target_supplier):
            continue
        result["review_status"] = status
        if status == "confirmed":
            result["user_confirmed"] = True
            if result.get("data"):
                result["data"]["needs_review"] = []
        changed += 1
    return changed


def _unique_success_suppliers(batch: dict) -> list[str]:
    suppliers = []
    for result in batch.get("results", []):
        if result.get("review_status") == "skipped":
            continue
        if result.get("status") != "success" or not result.get("data"):
            continue
        supplier = _result_supplier(result)
        if supplier and supplier != "未知供应商" and supplier not in suppliers:
            suppliers.append(supplier)
    return suppliers


def _resolve_action_supplier(batch: dict, target_supplier: str | None,
                             correction: dict) -> str | None:
    supplier = (target_supplier or correction.get("supplier") or "").strip()
    if supplier:
        return supplier
    suppliers = _unique_success_suppliers(batch)
    if len(suppliers) == 1:
        return suppliers[0]
    batch["needs_clarification"] = {
        "reason": "ambiguous_target",
        "source_text": correction.get("source_text", ""),
        "message": "这句话不知道要改哪一张，请说明供应商或单号。",
        "candidate_suppliers": suppliers,
    }
    return None


def _update_batch_summary(batch: dict):
    results = batch.get("results", [])
    confirmed = sum(1 for r in results if r.get("review_status") == "confirmed")
    pending = sum(1 for r in results if r.get("review_status") == "pending")
    skipped = sum(1 for r in results if r.get("review_status") == "skipped")
    batch["summary"] = {
        **batch.get("summary", {}),
        "total_invoices": len(results),
        "confirmed": confirmed,
        "pending_review": pending,
        "skipped": skipped,
    }


def _field_label_for_issue(field: str) -> str:
    labels = {
        "style_number": "款号",
        "material_type": "面料/辅料",
        "supplier_name": "面料厂家",
        "fabric_code": "面料款号",
        "color_code": "色号/颜色",
        "material_name": "品名",
        "unit_price": "单价",
        "quantity": "数量",
        "unit": "单位",
        "total_amount": "金额",
    }
    return labels.get(field, field)


def _remove_field_issue(result: dict, row_number: int, field: str):
    label = _field_label_for_issue(field)
    prefix = f"第{row_number}行 {label} "
    needs_review = result.get("data", {}).get("needs_review", [])
    result["data"]["needs_review"] = [
        issue for issue in needs_review
        if not str(issue).startswith(prefix)
    ]


def _refresh_row_update_review_status(batch: dict, correction: dict, target_supplier: str):
    row_number = int(correction.get("row_number") or 0)
    field = correction.get("field") or ""
    for result in batch.get("results", []):
        if result.get("status") != "success" or not result.get("data"):
            continue
        if not _result_matches_supplier(result, target_supplier):
            continue
        _remove_field_issue(result, row_number, field)
        if result["data"].get("needs_review"):
            result["review_status"] = "pending"
        else:
            result["review_status"] = "confirmed"
            result["user_confirmed"] = True


def _save_confirmed_templates(results_json: str, templates_dir: Path,
                              target_supplier: str | None,
                              corrections: list[dict]):
    corrections_json = (
        json.dumps(corrections, ensure_ascii=False)
        if corrections else None
    )
    save_from_results(
        results_json,
        None,
        templates_dir,
        corrections_json,
        target_supplier=target_supplier,
    )


def apply_finalize_actions(batch: dict, results_json: str, templates_dir: Path,
                           correction_texts: list[str], target_supplier: str | None) -> int:
    """Apply Feishu/OpenClaw-style confirmation and correction texts."""
    corrections = []
    for text in correction_texts:
        parsed = parse_natural_correction(text, target_supplier or "")
        for correction in parsed:
            correction.setdefault("source_text", text)
        corrections.extend(parsed)

    if not corrections:
        return 0

    changed = 0
    confirm_targets: set[str] = set()
    rule_targets: set[str] = set()
    learning_by_supplier: dict[str, list[dict]] = {}
    should_confirm_all = False

    for correction in corrections:
        action = correction.get("action")
        if action == "confirm_all":
            should_confirm_all = True
            continue

        supplier = _resolve_action_supplier(batch, target_supplier, correction)
        if not supplier:
            continue

        learning_by_supplier.setdefault(supplier, []).append(correction)

        if action == "confirm":
            confirm_targets.add(supplier)
        elif action == "skip_invoice":
            changed += _mark_matching_results(batch, supplier, "skipped")
        elif action == "row_update":
            changed += _apply_row_update(batch, correction, supplier)
            confirm_targets.add(supplier)
        elif correction.get("field") and correction.get("actual_meaning"):
            rule_targets.add(supplier)

    if should_confirm_all:
        changed += _mark_matching_results(batch, None, "confirmed")
        for supplier in _unique_success_suppliers(batch):
            learning_by_supplier.setdefault(supplier, []).extend(
                c for c in corrections if c.get("action") == "confirm_all"
            )

    for supplier, supplier_corrections in learning_by_supplier.items():
        _save_confirmed_templates(
            results_json,
            templates_dir,
            supplier,
            supplier_corrections,
        )

    for supplier in rule_targets:
        _events, rules = build_learning_entries(learning_by_supplier.get(supplier, []))
        learned = [rule["summary"] for rule in rules if rule.get("summary")]
        for result in batch.get("results", []):
            if result.get("review_status") == "skipped":
                continue
            if result.get("status") != "success" or not result.get("data"):
                continue
            if _result_matches_supplier(result, supplier):
                result["learned"] = learned
                post_process_extraction(result, templates_dir)

    for correction in corrections:
        if correction.get("action") != "row_update":
            continue
        supplier = _resolve_action_supplier(batch, target_supplier, correction)
        if supplier:
            _refresh_row_update_review_status(batch, correction, supplier)

    for supplier in confirm_targets:
        has_row_update = any(
            correction.get("action") == "row_update"
            and _resolve_action_supplier(batch, target_supplier, correction) == supplier
            for correction in corrections
        )
        has_explicit_confirm = any(
            correction.get("action") == "confirm"
            and _resolve_action_supplier(batch, target_supplier, correction) == supplier
            for correction in corrections
        )
        if has_row_update and not has_explicit_confirm:
            continue
        changed += _mark_matching_results(batch, supplier, "confirmed")

    _update_batch_summary(batch)
    with open(results_json, "w", encoding="utf-8") as f:
        json.dump(batch, f, ensure_ascii=False, indent=2)

    return changed


def check_existing_templates(templates_dir: Path) -> bool:
    """Check if templates exist and print status. Returns True if templates found."""
    index_path = templates_dir / "index.json"
    if not index_path.is_file():
        return False
    try:
        with open(index_path, "r", encoding="utf-8") as f:
            index = json.load(f)
    except (json.JSONDecodeError, OSError):
        return False
    templates = index.get("templates", {})
    if not templates:
        return False
    print(f"EXISTING_TEMPLATES:{len(templates)}")
    for sid, entry in templates.items():
        print(f"  - {entry.get('supplier_name', sid)}")
    return True


def count_existing_templates(templates_dir: Path) -> int:
    index_path = templates_dir / "index.json"
    if not index_path.is_file():
        return 0
    try:
        with open(index_path, "r", encoding="utf-8") as f:
            return len(json.load(f).get("templates", {}))
    except (json.JSONDecodeError, OSError):
        return 0


def _resolve_gateway_config(args) -> tuple:
    """Resolve gateway host, port, and token from args, env, or OpenClaw config."""
    host = args.gateway_host
    port = args.gateway_port
    token = args.token

    # If token is provided via args or env, use as-is
    if token:
        return host, port, token

    # Try reading from OpenClaw config
    config_paths = [
        Path.home() / ".openclaw" / "openclaw.json",
        Path.home() / ".openclaw" / "config.yaml",
        Path.home() / ".openclaw" / "gateway.yaml",
    ]
    for cfg_path in config_paths:
        if not cfg_path.is_file():
            continue
        try:
            if cfg_path.suffix == ".json":
                with open(cfg_path) as f:
                    cfg = json.load(f)
                gw = cfg.get("gateway", {})
                auth = gw.get("auth", {})
                if auth.get("mode") == "token":
                    token = auth.get("token", "")
                elif auth.get("mode") == "password":
                    token = auth.get("password", "")
                if not host or host == "127.0.0.1":
                    host = gw.get("http", {}).get("host", "127.0.0.1")
                if port == 18789:
                    port = gw.get("http", {}).get("port", 18789)
                if token:
                    break
        except (json.JSONDecodeError, OSError, KeyError):
            continue

    if not token:
        print("Warning: No gateway token found. Set OPENCLAW_GATEWAY_TOKEN or configure gateway auth.", file=sys.stderr)

    return host, port, token



def main():
    parser = argparse.ArgumentParser(
        description="Invoice OCR unified pipeline: extract → review → export")
    parser.add_argument("input_dir", nargs="?", help="Directory containing invoice images")
    parser.add_argument("--intro", action="store_true",
                        help="Print first-use introduction and exit")
    parser.add_argument("--self-check", action="store_true",
                        help="Check whether the current OpenClaw model can see images")
    parser.add_argument("--output", "-o",
                        help="Output Excel file path (if not set, uses saved preference or asks user)")
    parser.add_argument("--finalize", action="store_true",
                        help="Finalize: apply pending corrections and export all")
    parser.add_argument("--corrections", nargs="*", default=[],
                        help="Natural language corrections (e.g., '品名就是面料款号')")
    parser.add_argument("--supplier", type=str,
                        help="Apply corrections to specific supplier")
    parser.add_argument("--gateway-host",
                        default=os.environ.get("OPENCLAW_GATEWAY_HOST", DEFAULT_GATEWAY_HOST))
    parser.add_argument("--gateway-port", type=int,
                        default=int(os.environ.get("OPENCLAW_GATEWAY_PORT", str(DEFAULT_GATEWAY_PORT))))
    parser.add_argument("--token", default=os.environ.get("OPENCLAW_GATEWAY_TOKEN", ""))
    parser.add_argument("--model", default="openclaw/default")
    parser.add_argument("--parallel", "-j", type=int, default=3)
    parser.add_argument("--retries", "-r", type=int, default=MAX_RETRIES)
    parser.add_argument("--templates-dir", type=str, default=str(TEMPLATES_DIR))
    parser.add_argument("--append", action="store_true",
                        help="Append to existing results")
    parser.add_argument("--review-in-excel", action="store_true",
                        help="Export successful pending rows to Excel for manual review")
    parser.add_argument("--assistant-summary-json", action="store_true",
                        help="Print one channel-friendly summary JSON line for OpenClaw/Feishu")
    parser.add_argument("--agent-mode", action="store_true",
                        help="Single-command mode for Agent: process + show results in one JSON line, no mixed stdout")
    parser.add_argument("--apply", type=str,
                        help="Apply corrections in one shot. Format: 'Supplier:correction | Supplier:correction' or '全部确认'")
    parser.add_argument("--interactive", action="store_true",
                        help="Interactive mode: walk through pending items one by one in terminal")
    parser.add_argument("--setup-table",
                        help="Validate or create the target Excel table")
    parser.add_argument("--headers",
                        help="Comma-separated headers used with --setup-table")
    parser.add_argument("--header-mapping",
                        help="JSON object: 1-based column index to allowed field or null")
    args = parser.parse_args()

    # Auto-detect gateway config if not explicitly provided
    args.gateway_host, args.gateway_port, args.token = _resolve_gateway_config(args)

    # --agent-mode: suppress extraneous stdout, act as final mode
    if args.agent_mode:
        args.assistant_summary_json = True

    templates_dir = Path(args.templates_dir)

    prefs = load_preferences()
    saved_table = prefs.get("default_table")
    target_table_decision = None

    if args.self_check:
        base_url = f"http://{args.gateway_host}:{args.gateway_port}"
        try:
            check_gateway(base_url, args.token)
        except SystemExit:
            check = {"vision_supported": False, "reason": "gateway_error"}
            summary = build_self_check_summary(check)
            if args.assistant_summary_json:
                print_assistant_summary(summary, clean=args.agent_mode)
                sys.exit(0)
            print(summary["user_message"])
            sys.exit(1)
        check = check_model_vision_support(base_url, args.token, args.model)
        summary = build_self_check_summary(check)
        if args.assistant_summary_json:
            print_assistant_summary(summary, clean=args.agent_mode)
        else:
            print(summary["user_message"])
        sys.exit(0)

    if args.intro:
        template_count = count_existing_templates(templates_dir)
        summary = build_intro_summary(bool(saved_table), template_count)
        if args.assistant_summary_json:
            print_assistant_summary(summary, clean=args.agent_mode)
        else:
            print(summary["user_message"])
        sys.exit(0)

    if args.setup_table:
        table_path = args.setup_table
        if not Path(table_path).is_file():
            if not args.headers:
                parser.error("--headers is required when creating a new table")
            create_target_table(table_path, args.headers)
        if args.header_mapping:
            try:
                user_mapping = json.loads(args.header_mapping)
            except json.JSONDecodeError as error:
                parser.error(f"invalid --header-mapping JSON: {error}")
            if not isinstance(user_mapping, dict):
                parser.error("--header-mapping must be a JSON object")
            save_user_header_mapping(table_path, user_mapping)

        target_table_decision = prepare_target_table(
            table_path,
            None,
            {
                "base_url": (
                    f"http://{args.gateway_host}:{args.gateway_port}"
                ),
                "token": args.token,
                "model": args.model,
            },
        )
        if target_table_decision["status"] != "needs_header_confirmation":
            set_default_table(table_path)
            saved_table = table_path

        if not args.input_dir:
            if target_table_decision["status"] == "needs_header_confirmation":
                summary = {
                    "status": "needs_header_confirmation",
                    "user_message": (
                        "这张表有几列表头还不能确定，请确认字段对应关系。"
                    ),
                    "pending_mappings": target_table_decision["pending_mappings"],
                }
            else:
                summary = {
                    "status": "table_ready",
                    "user_message": (
                        f"已设置录入表《{Path(table_path).name}》，"
                        "以后会默认继续录入到这张表。"
                    ),
                    "pending_mappings": target_table_decision["pending_mappings"],
                }
            if args.assistant_summary_json:
                print_assistant_summary(summary, clean=args.agent_mode)
            else:
                print(summary["user_message"])
            sys.exit(0)

    if not args.input_dir:
        parser.error("input_dir is required unless --intro is used")

    # First-run check: detect existing templates and preferences
    if not args.agent_mode:
        check_existing_templates(templates_dir)
    if saved_table:
        if not args.agent_mode:
            print(f"SAVED_TABLE:{saved_table}")

    # Resolve output path: explicit arg > saved preference > ask
    output = args.output
    if not output:
        output = saved_table
    if not output:
        if not args.agent_mode:
            print("NO_TABLE_SPECIFIED")
        if args.assistant_summary_json:
            if args.agent_mode:
                print(json.dumps(build_need_table_summary(), ensure_ascii=False, separators=(',', ':')))
            else:
                print_assistant_summary(build_need_table_summary())
        sys.exit(0)

    if target_table_decision is None:
        target_table_decision = prepare_target_table(
            output,
            saved_table,
            {
                "base_url": f"http://{args.gateway_host}:{args.gateway_port}",
                "token": args.token,
                "model": args.model,
            },
        )
    if target_table_decision["status"] == "needs_header_confirmation":
        summary = {
            "status": "needs_header_confirmation",
            "user_message": "这张表有几列表头还不能确定，请确认字段对应关系。",
            "pending_mappings": target_table_decision["pending_mappings"],
        }
        if args.assistant_summary_json:
            print_assistant_summary(summary, clean=args.agent_mode)
        else:
            print(summary["user_message"])
        sys.exit(0)
    if target_table_decision["status"] == "waiting_for_table":
        if args.assistant_summary_json:
            print_assistant_summary(build_need_table_summary(), clean=args.agent_mode)
        else:
            print("NO_TABLE_SPECIFIED")
        sys.exit(0)
    table_mappings = target_table_decision.get("mappings") or None
    pending_header_mappings = target_table_decision.get("pending_mappings") or []

    # Handle --apply: shortcut for corrections + finalize in one step
    if args.apply or args.finalize:
        results_json = str(Path(args.input_dir) / "results.json")

        # If --apply was used and results.json doesn't exist yet, process first
        if args.apply and not Path(results_json).is_file():
            result = process_batch(
                args.input_dir, output,
                args.gateway_host, args.gateway_port, args.token,
                args.model, templates_dir,
                args.parallel, args.retries, args.append, args.review_in_excel,
                args.assistant_summary_json,
                table_mappings,
                pending_header_mappings,
            )
            # Write results.json from in-memory batch
            if result.get("batch"):
                os.makedirs(Path(results_json).parent, exist_ok=True)
                with open(results_json, "w", encoding="utf-8") as f:
                    json.dump(result["batch"], f, ensure_ascii=False, indent=2)

        if not Path(results_json).is_file():
            msg = "Error: No results found. Run without --apply first, or provide images."
            if args.agent_mode:
                print(json.dumps({"status": "error", "user_message": msg}, ensure_ascii=False))
            else:
                print(msg, file=sys.stderr)
            sys.exit(1)

        with open(results_json, "r", encoding="utf-8") as f:
            batch = json.load(f)

        # Parse --apply string into corrections list
        all_corrections = list(args.corrections)
        if args.apply:
            # Format: "旺泰:确认 | 宇博:无色号" or just "全部确认"
            parts = [p.strip() for p in args.apply.split("|")]
            for part in parts:
                if ":" in part:
                    supplier, correction = part.split(":", 1)
                    all_corrections.append(correction.strip())
                else:
                    all_corrections.append(part)

        if all_corrections:
            changed = apply_finalize_actions(
                batch,
                results_json,
                templates_dir,
                all_corrections,
                args.supplier,
            )
            if not args.agent_mode:
                print(f"Applied {changed} update(s).")
            if batch.get("needs_clarification"):
                summary_result = build_result_view(batch, output)
                summary_result["pending_mappings"] = pending_header_mappings
                summary_result["needs_clarification"] = batch["needs_clarification"]
                if args.agent_mode:
                    print(json.dumps(build_assistant_summary(summary_result), ensure_ascii=False, separators=(',', ':')))
                elif args.assistant_summary_json:
                    print_assistant_summary(build_assistant_summary(summary_result))
                else:
                    print(batch["needs_clarification"].get("message", "请说明要修改哪一张票据。"))
                sys.exit(0)

        # Export
        export_xlsx_append(batch, output, table_mappings)
        total = sum(1 for r in batch.get("results", []) if r.get("review_status") == "confirmed")
        pending = sum(1 for r in batch.get("results", []) if r.get("review_status") == "pending")
        if not args.agent_mode:
            print(f"Exported to {output} ({total} confirmed, {pending} still pending)")
        summary_result = build_result_view(batch, output)
        summary_result["pending_mappings"] = pending_header_mappings
        if args.agent_mode:
            print(json.dumps(build_assistant_summary(summary_result), ensure_ascii=False, separators=(',', ':')))
        elif args.assistant_summary_json:
            print_assistant_summary(build_assistant_summary(summary_result))

        set_default_table(output)
    else:
        result = process_batch(
            args.input_dir, output,
            args.gateway_host, args.gateway_port, args.token,
            args.model, templates_dir,
            args.parallel, args.retries, args.append, args.review_in_excel,
            args.assistant_summary_json,
            table_mappings,
            pending_header_mappings,
        )

        # Interactive mode: walk through pending items
        if args.interactive and result.get("batch") and result.get("pending"):
            results_json = str(Path(args.input_dir) / "results.json")
            if not Path(results_json).is_file():
                os.makedirs(Path(results_json).parent, exist_ok=True)
                with open(results_json, "w", encoding="utf-8") as f:
                    json.dump(result["batch"], f, ensure_ascii=False, indent=2)
            _run_interactive_loop(result["batch"], output, results_json, templates_dir)
            # Export after interactive confirmations
            export_xlsx_append(result["batch"], output, table_mappings)
            set_default_table(output)
            summary_result = build_result_view(result["batch"], output)
            summary_result["pending_mappings"] = pending_header_mappings
            summary = build_assistant_summary(summary_result)
            print()
            print(summary["user_message"])
        else:
            if not args.agent_mode:
                print_summary(result, templates_dir)
            if args.assistant_summary_json:
                summary = build_assistant_summary(result)
                if args.agent_mode:
                    print(json.dumps(summary, ensure_ascii=False, separators=(',', ':')))
                else:
                    print_assistant_summary(summary)


if __name__ == "__main__":
    main()
